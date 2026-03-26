import sys
import os
import time

def write_prog(progress_file, prog):
    if progress_file:
        try:
            with open(progress_file, 'w') as f:
                f.write(str(prog))
        except:
            pass

def excel_to_csv(input_path, output_path, progress_file=None):
    try:
        import openpyxl
        import csv
        
        write_prog(progress_file, 10)
        
        # Read Excel
        wb = openpyxl.load_workbook(input_path, data_only=True)
        sheet = wb.active
        
        write_prog(progress_file, 50)
        
        # Write CSV
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            c = csv.writer(f)
            for r in sheet.iter_rows(values_only=True):
                c.writerow(r)
                
        write_prog(progress_file, 100)
        
    except Exception as e:
        sys.stderr.write(f"Excel to CSV failed: {str(e)}\n")
        sys.exit(1)

def csv_to_excel(input_path, output_path, progress_file=None):
    try:
        import openpyxl
        import csv
        
        write_prog(progress_file, 10)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        write_prog(progress_file, 30)
        
        # Read CSV and Write Excel
        with open(input_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:
                ws.append(row)
                
        write_prog(progress_file, 70)
        
        wb.save(output_path)
        write_prog(progress_file, 100)
        
    except Exception as e:
        sys.stderr.write(f"CSV to Excel failed: {str(e)}\n")
        sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) < 4:
        sys.stderr.write("Usage: python excel_convert.py <input> <output> <--to-csv|--to-excel> [progress.txt]\n")
        sys.exit(1)
        
    input_path = sys.argv[1]
    output_path = sys.argv[2]
    mode_flag = sys.argv[3]
    progress_file = sys.argv[4] if len(sys.argv) > 4 else None
    
    if not os.path.exists(input_path):
        sys.stderr.write(f"Error: input file {input_path} not found.\n")
        sys.exit(1)
    
    if mode_flag == "--to-csv":
        excel_to_csv(input_path, output_path, progress_file)
    elif mode_flag == "--to-excel":
        csv_to_excel(input_path, output_path, progress_file)
    else:
        sys.stderr.write(f"Error: Unknown flag {mode_flag}\n")
        sys.exit(1)
    
    if not os.path.exists(output_path):
        sys.stderr.write("Error: Output file was not generated.\n")
        sys.exit(1)
